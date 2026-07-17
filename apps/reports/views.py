import io
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from apps.exams.models import Ujian
from apps.submissions.models import SesiUjian, Jawaban
from apps.proctoring.models import PelanggaranLog


def require_dosen(func):
    def wrapper(self, request, *args, **kwargs):
        if not request.user.is_dosen:
            return Response({'detail': 'Akses ditolak. Hanya dosen.'}, status=403)
        return func(self, request, *args, **kwargs)
    return wrapper


class NilaiUjianView(APIView):
    """
    GET /api/v1/laporan/nilai/<ujian_pk>/?kelas=TI-3A
    Tabel nilai semua peserta satu ujian.
    """
    permission_classes = [IsAuthenticated]

    @require_dosen
    def get(self, request, ujian_pk):
        ujian = get_object_or_404(Ujian, pk=ujian_pk, mata_pelajaran__dosen=request.user)
        kelas_filter = request.query_params.get('kelas', '')
        qs = SesiUjian.objects.filter(ujian=ujian).select_related('mahasiswa').prefetch_related(
            'jawaban__soal'
        ).order_by('mahasiswa__kelas', 'mahasiswa__nama_lengkap')
        if kelas_filter:
            qs = qs.filter(mahasiswa__kelas=kelas_filter)

        soal_list = list(ujian.soal.order_by('nomor_urut'))
        kelas_list = sorted(set(
            s.mahasiswa.kelas for s in SesiUjian.objects.filter(ujian=ujian).select_related('mahasiswa')
        ))

        peserta = []
        for sesi in qs:
            jawaban_map = {j.soal_id: j for j in sesi.jawaban.all()}
            nilai_per_soal = {}
            for soal in soal_list:
                j = jawaban_map.get(soal.pk)
                nilai_per_soal[f"soal_{soal.nomor_urut}"] = j.nilai if j and j.nilai is not None else None
            peserta.append({
                'sesi_id': sesi.pk,
                'nama': sesi.mahasiswa.nama_lengkap,
                'nim': sesi.mahasiswa.nim,
                'kelas': sesi.mahasiswa.kelas,
                'status': sesi.status,
                'total_nilai': sesi.total_nilai,
                'nilai_maksimal': ujian.nilai_maksimal,
                'nilai_per_soal': nilai_per_soal,
            })

        return Response({
            'ujian': {'id': ujian.pk, 'judul': ujian.judul, 'kode': ujian.mata_pelajaran.kode},
            'kelas_list': kelas_list,
            'kelas_filter': kelas_filter,
            'peserta': peserta,
        })


class LogPelanggaranView(APIView):
    """GET /api/v1/laporan/log-pelanggaran/<ujian_pk>/"""
    permission_classes = [IsAuthenticated]

    @require_dosen
    def get(self, request, ujian_pk):
        ujian = get_object_or_404(Ujian, pk=ujian_pk, mata_pelajaran__dosen=request.user)
        logs = PelanggaranLog.objects.filter(
            sesi__ujian=ujian
        ).select_related('sesi__mahasiswa').order_by('-timestamp')
        return Response([{
            'mahasiswa': log.sesi.mahasiswa.nama_lengkap,
            'nim': log.sesi.mahasiswa.nim,
            'tipe': log.tipe,
            'tipe_display': log.get_tipe_display(),
            'timestamp': log.timestamp,
            'keterangan': log.keterangan,
        } for log in logs])


class ExportExcelView(APIView):
    """GET /api/v1/laporan/export/excel/<ujian_pk>/?kelas=TI-3A"""
    permission_classes = [IsAuthenticated]

    @require_dosen
    def get(self, request, ujian_pk):
        ujian = get_object_or_404(Ujian, pk=ujian_pk, mata_pelajaran__dosen=request.user)
        kelas_filter = request.query_params.get('kelas', '')
        qs = SesiUjian.objects.filter(ujian=ujian).select_related('mahasiswa').prefetch_related(
            'jawaban__soal'
        ).order_by('mahasiswa__kelas', 'mahasiswa__nama_lengkap')
        if kelas_filter:
            qs = qs.filter(mahasiswa__kelas=kelas_filter)

        soal_list = list(ujian.soal.order_by('nomor_urut'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nilai Ujian"

        hdr_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alt_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")

        headers = ['No', 'Nama Lengkap', 'NIM', 'Kelas']
        for s in soal_list:
            headers.append(f'Soal {s.nomor_urut}')
        headers += ['Total Nilai', 'Nilai Maks', 'Status']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin

        for row_num, sesi in enumerate(qs, 2):
            fill = alt_fill if row_num % 2 == 0 else PatternFill()
            jmap = {j.soal_id: j for j in sesi.jawaban.all()}
            row_data = [row_num - 1, sesi.mahasiswa.nama_lengkap, sesi.mahasiswa.nim, sesi.mahasiswa.kelas]
            for soal in soal_list:
                j = jmap.get(soal.pk)
                row_data.append(j.nilai if j and j.nilai is not None else '-')
            row_data += [
                sesi.total_nilai if sesi.total_nilai is not None else '-',
                ujian.nilai_maksimal,
                sesi.get_status_display(),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.alignment = center
                cell.border = thin
                cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"laporan_{ujian.mata_pelajaran.kode}_{kelas_filter or 'semua'}.xlsx"
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response


class ExportPDFMahasiswaView(APIView):
    """GET /api/v1/laporan/export/pdf/<sesi_pk>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, sesi_pk):
        if request.user.is_mahasiswa:
            sesi = get_object_or_404(SesiUjian, pk=sesi_pk, mahasiswa=request.user)
        else:
            sesi = get_object_or_404(SesiUjian, pk=sesi_pk, ujian__mata_pelajaran__dosen=request.user)

        jawaban_list = sesi.jawaban.all().select_related('soal').order_by('soal__nomor_urut')
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                                leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        el = []

        title_s = ParagraphStyle('t', parent=styles['Heading1'], alignment=1,
                                 textColor=colors.HexColor('#1e3a5f'), spaceAfter=8)
        sub_s = ParagraphStyle('s', parent=styles['Normal'], alignment=1,
                               textColor=colors.HexColor('#555'), spaceAfter=4)
        el.append(Paragraph("LAPORAN HASIL UJIAN ESAI", title_s))
        el.append(Paragraph(sesi.ujian.judul, sub_s))
        el.append(Paragraph(sesi.ujian.mata_pelajaran.nama, sub_s))
        el.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3a5f')))
        el.append(Spacer(1, 0.4*cm))

        info = [
            ['Nama', ':', sesi.mahasiswa.nama_lengkap],
            ['NIM', ':', sesi.mahasiswa.nim],
            ['Kelas', ':', sesi.mahasiswa.kelas],
            ['Waktu Mulai', ':', sesi.waktu_mulai.strftime('%d/%m/%Y %H:%M') if sesi.waktu_mulai else '-'],
            ['Waktu Selesai', ':', sesi.waktu_selesai.strftime('%d/%m/%Y %H:%M') if sesi.waktu_selesai else '-'],
            ['Total Nilai', ':', f"{sesi.total_nilai} / {sesi.ujian.nilai_maksimal}" if sesi.total_nilai is not None else 'Menunggu penilaian'],
        ]
        info_t = Table(info, colWidths=[3.5*cm, 0.5*cm, 12*cm])
        info_t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        el.append(info_t)
        el.append(Spacer(1, 0.5*cm))
        el.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#ccc')))
        el.append(Spacer(1, 0.3*cm))
        el.append(Paragraph("Detail Penilaian Per Soal", styles['Heading2']))
        el.append(Spacer(1, 0.3*cm))

        for j in jawaban_list:
            soal_s = ParagraphStyle('q', parent=styles['Normal'],
                                    fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a5f'))
            el.append(Paragraph(f"Soal {j.soal.nomor_urut}: {j.soal.pertanyaan}", soal_s))
            el.append(Spacer(1, 0.15*cm))
            el.append(Paragraph(f"<b>Jawaban:</b> {j.teks_jawaban or '(Tidak dijawab)'}", styles['Normal']))
            nc = colors.HexColor('#2ecc71') if j.nilai == 10 else (colors.HexColor('#f39c12') if j.nilai == 5 else colors.HexColor('#e74c3c'))
            ns = ParagraphStyle('n', parent=styles['Normal'], textColor=nc, fontName='Helvetica-Bold')
            el.append(Paragraph(f"Nilai: {j.nilai if j.nilai is not None else 'Menunggu'}", ns))
            if j.alasan_nilai:
                el.append(Paragraph(f"<i>Catatan AI: {j.alasan_nilai}</i>", styles['Normal']))
            el.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#eee')))
            el.append(Spacer(1, 0.3*cm))

        doc.build(el)
        buf.seek(0)
        fname = f"hasil_{sesi.mahasiswa.nim}_{sesi.ujian.mata_pelajaran.kode}.pdf"
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response
